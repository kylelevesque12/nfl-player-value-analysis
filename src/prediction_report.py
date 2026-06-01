"""Utilities for creating 2026 NFL player value predictions.

The report uses the enhanced-history Random Forest approach from Notebook 05.
It trains on historical player-season rows where the next-season value score
is known, then applies the model to 2025 player seasons to create 2026
projections and confidence labels.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.impute import SimpleImputer
from sklearn.metrics import (
    accuracy_score,
    brier_score_loss,
    mean_absolute_error,
    mean_squared_error,
    r2_score,
    roc_auc_score,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler

from src import config

RAW_PRODUCTION_FEATURES = [
    "position",
    "age",
    "years_exp",
    "draft_number",
    "games_played",
    "value_epa_total",
    "value_epa_per_game",
    "yards_per_game",
    "tds_per_game",
]

HISTORY_FEATURES = [
    "prior_qualifying_seasons",
    "value_score_prev",
    "value_score_last2_avg",
    "value_score_last3_avg",
    "value_score_trend_2yr",
    "value_epa_total_prev",
    "value_epa_per_game_prev",
    "games_played_prev",
    "games_played_last2_sum",
    "games_played_last3_avg",
    "yards_per_game_prev",
    "tds_per_game_prev",
]

ENHANCED_FEATURES = RAW_PRODUCTION_FEATURES + HISTORY_FEATURES
MIN_VALUE_GAMES = config.MIN_VALUE_GAMES
PREDICTION_INTERVAL_MULTIPLIER = config.PREDICTION_INTERVAL_MULTIPLIER
PREDICTION_INTERVAL_TARGET_COVERAGE = config.PREDICTION_INTERVAL_TARGET_COVERAGE
CSV_FLOAT_FORMAT = config.CSV_FLOAT_FORMAT

TUNED_RANDOM_FOREST_PARAMS = {
    "n_estimators": 300,
    "max_depth": 7,
    "max_features": 0.5,
    "min_samples_leaf": 20,
    "random_state": 42,
    "n_jobs": -1,
}

AVAILABILITY_RANDOM_FOREST_PARAMS = {
    "n_estimators": 300,
    "max_depth": 6,
    "max_features": 0.75,
    "min_samples_leaf": 10,
    "class_weight": "balanced_subsample",
    "random_state": 42,
    "n_jobs": -1,
}

SUM_COLS = [
    "completions",
    "attempts",
    "passing_yards",
    "passing_tds",
    "passing_interceptions",
    "sacks_suffered",
    "sack_yards_lost",
    "passing_air_yards",
    "passing_yards_after_catch",
    "passing_first_downs",
    "passing_epa",
    "passing_2pt_conversions",
    "carries",
    "rushing_yards",
    "rushing_tds",
    "rushing_first_downs",
    "rushing_epa",
    "rushing_2pt_conversions",
    "targets",
    "receptions",
    "receiving_yards",
    "receiving_tds",
    "receiving_air_yards",
    "receiving_yards_after_catch",
    "receiving_first_downs",
    "receiving_epa",
    "receiving_2pt_conversions",
    "fantasy_points",
    "fantasy_points_ppr",
    "games_played",
]

FIRST_COLS = [
    "player_name",
    "player_display_name",
    "position",
    "birth_date",
    "height",
    "weight",
    "years_exp",
    "entry_year",
    "rookie_year",
    "draft_club",
    "draft_number",
    "college",
    "age",
]


def find_project_root(expected_file: str = "data/processed/player_value_scores_2016_2025.csv") -> Path:
    """Find the project root from common terminal or notebook working directories."""
    for candidate in [Path.cwd(), *Path.cwd().parents]:
        if (candidate / expected_file).exists():
            return candidate
    raise FileNotFoundError(
        "Could not find " + expected_file + " from working directory " + str(Path.cwd())
    )


def _series_or_zero(df: pd.DataFrame, col: str) -> pd.Series:
    if col in df.columns:
        return pd.to_numeric(df[col], errors="coerce").fillna(0)
    return pd.Series(0, index=df.index, dtype="float64")


def add_group_zscore(
    df: pd.DataFrame,
    value_col: str,
    z_col: str,
    group_cols: list[str] | None = None,
) -> pd.DataFrame:
    """Add a z-score column within season-position groups."""
    if group_cols is None:
        group_cols = ["season", "position"]

    df = df.copy()
    group_mean = df.groupby(group_cols)[value_col].transform("mean")
    group_std = df.groupby(group_cols)[value_col].transform("std")
    df[z_col] = (df[value_col] - group_mean) / group_std
    return df


def collapse_to_player_season(value_scored: pd.DataFrame) -> pd.DataFrame:
    """Collapse player-team rows to one row per player-season for modeling."""
    value_scored = value_scored.copy()

    available_sum_cols = [col for col in SUM_COLS if col in value_scored.columns]
    available_first_cols = [col for col in FIRST_COLS if col in value_scored.columns]

    team_sort_cols = ["season", "player_id"]
    if "games_played" in value_scored.columns:
        team_sort_cols.append("games_played")
    if "value_epa_total" in value_scored.columns:
        team_sort_cols.append("value_epa_total")

    ascending = [True, True] + [False] * (len(team_sort_cols) - 2)
    primary_team = (
        value_scored.sort_values(team_sort_cols, ascending=ascending)
        .drop_duplicates(["season", "player_id"])[["season", "player_id", "team"]]
        .rename(columns={"team": "primary_team"})
    )

    agg_spec: dict[str, Any] = {col: "sum" for col in available_sum_cols}
    agg_spec.update({col: "first" for col in available_first_cols})
    agg_spec["team"] = lambda x: "/".join(sorted(x.dropna().astype(str).unique()))

    player_season = (
        value_scored.groupby(["season", "player_id"], as_index=False).agg(agg_spec)
    )
    player_season = player_season.rename(columns={"team": "teams"})
    player_season = player_season.merge(primary_team, on=["season", "player_id"], how="left")

    player_season["total_yards"] = (
        _series_or_zero(player_season, "passing_yards")
        + _series_or_zero(player_season, "rushing_yards")
        + _series_or_zero(player_season, "receiving_yards")
    )
    player_season["total_tds"] = (
        _series_or_zero(player_season, "passing_tds")
        + _series_or_zero(player_season, "rushing_tds")
        + _series_or_zero(player_season, "receiving_tds")
    )
    player_season["total_epa"] = (
        _series_or_zero(player_season, "passing_epa")
        + _series_or_zero(player_season, "rushing_epa")
        + _series_or_zero(player_season, "receiving_epa")
    )
    player_season["yards_per_game"] = player_season["total_yards"] / player_season["games_played"].replace(0, np.nan)
    player_season["tds_per_game"] = player_season["total_tds"] / player_season["games_played"].replace(0, np.nan)
    player_season["epa_per_game"] = player_season["total_epa"] / player_season["games_played"].replace(0, np.nan)

    player_season["scrimmage_touches"] = _series_or_zero(player_season, "carries") + _series_or_zero(player_season, "receptions")
    player_season["scrimmage_yards"] = _series_or_zero(player_season, "rushing_yards") + _series_or_zero(player_season, "receiving_yards")
    player_season["scrimmage_tds"] = _series_or_zero(player_season, "rushing_tds") + _series_or_zero(player_season, "receiving_tds")
    player_season["scrimmage_epa"] = _series_or_zero(player_season, "rushing_epa") + _series_or_zero(player_season, "receiving_epa")
    player_season["scrimmage_yards_per_game"] = player_season["scrimmage_yards"] / player_season["games_played"].replace(0, np.nan)
    player_season["scrimmage_touches_per_game"] = player_season["scrimmage_touches"] / player_season["games_played"].replace(0, np.nan)
    player_season["scrimmage_tds_per_game"] = player_season["scrimmage_tds"] / player_season["games_played"].replace(0, np.nan)
    player_season["scrimmage_epa_per_game"] = player_season["scrimmage_epa"] / player_season["games_played"].replace(0, np.nan)
    player_season["yards_per_scrimmage_touch"] = player_season["scrimmage_yards"] / player_season["scrimmage_touches"].replace(0, np.nan)

    player_season["qb_plays"] = _series_or_zero(player_season, "attempts") + _series_or_zero(player_season, "carries")
    player_season["qb_total_yards"] = _series_or_zero(player_season, "passing_yards") + _series_or_zero(player_season, "rushing_yards")
    player_season["qb_total_tds"] = _series_or_zero(player_season, "passing_tds") + _series_or_zero(player_season, "rushing_tds")
    player_season["qb_epa"] = _series_or_zero(player_season, "passing_epa") + _series_or_zero(player_season, "rushing_epa")
    player_season["qb_yards_per_play"] = player_season["qb_total_yards"] / player_season["qb_plays"].replace(0, np.nan)
    player_season["qb_yards_per_game"] = player_season["qb_total_yards"] / player_season["games_played"].replace(0, np.nan)
    player_season["qb_tds_per_game"] = player_season["qb_total_tds"] / player_season["games_played"].replace(0, np.nan)
    player_season["qb_epa_per_game"] = player_season["qb_epa"] / player_season["games_played"].replace(0, np.nan)
    player_season["interceptions_per_game"] = _series_or_zero(player_season, "passing_interceptions") / player_season["games_played"].replace(0, np.nan)

    player_season["value_epa_total"] = np.where(
        player_season["position"].eq("QB"),
        player_season["qb_epa"],
        player_season["scrimmage_epa"],
    )
    player_season["value_epa_per_game"] = player_season["value_epa_total"] / player_season["games_played"].replace(0, np.nan)

    player_season = add_group_zscore(player_season, "value_epa_total", "value_score")
    player_season = add_group_zscore(player_season, "value_epa_per_game", "value_score_per_game")
    player_season["value_score_gap"] = player_season["value_score_per_game"] - player_season["value_score"]

    return player_season


def create_player_season_value_scores(
    player_rows: pd.DataFrame,
    min_games: int = MIN_VALUE_GAMES,
) -> pd.DataFrame:
    """Create filtered player-season value scores after collapsing team stints."""
    player_season = collapse_to_player_season(player_rows)
    player_season = player_season[player_season["games_played"].ge(min_games)].copy()

    player_season = add_group_zscore(player_season, "value_epa_total", "value_score")
    player_season = add_group_zscore(player_season, "value_epa_per_game", "value_score_per_game")
    player_season["value_score_total_epa"] = player_season["value_score"]
    player_season["value_score_gap"] = player_season["value_score_per_game"] - player_season["value_score"]
    player_season["value_metric"] = "position_adjusted_total_epa"
    player_season["team"] = player_season["primary_team"]

    player_season["position_season_rank"] = (
        player_season
        .groupby(["season", "position"])["value_score"]
        .rank(ascending=False, method="min")
    )
    player_season["position_season_percentile"] = (
        player_season
        .groupby(["season", "position"])["value_score"]
        .rank(pct=True)
    )

    return player_season


def add_player_history_features(player_season: pd.DataFrame) -> pd.DataFrame:
    """Add lagged and rolling player-history features without future leakage."""
    player_season = player_season.sort_values(["player_id", "season"]).copy()
    grouped = player_season.groupby("player_id", group_keys=False)

    player_season["prior_qualifying_seasons"] = grouped.cumcount()

    history_cols = [
        "value_score",
        "value_epa_total",
        "value_epa_per_game",
        "games_played",
        "yards_per_game",
        "tds_per_game",
    ]

    for col in history_cols:
        if col not in player_season.columns:
            continue

        shifted = grouped[col].shift(1)
        player_season[f"{col}_prev"] = shifted
        player_season[f"{col}_last2_avg"] = grouped[col].apply(
            lambda s: s.shift(1).rolling(2, min_periods=1).mean()
        )
        player_season[f"{col}_last3_avg"] = grouped[col].apply(
            lambda s: s.shift(1).rolling(3, min_periods=1).mean()
        )

    if {"value_score_prev", "value_score_last2_avg"}.issubset(player_season.columns):
        player_season["value_score_trend_2yr"] = (
            player_season["value_score_prev"] - player_season["value_score_last2_avg"]
        )

    if "games_played" in player_season.columns:
        player_season["games_played_last2_sum"] = grouped["games_played"].apply(
            lambda s: s.shift(1).rolling(2, min_periods=1).sum()
        )
        player_season["games_played_last3_avg"] = grouped["games_played"].apply(
            lambda s: s.shift(1).rolling(3, min_periods=1).mean()
        )

    return player_season


def create_next_season_targets(player_season: pd.DataFrame) -> pd.DataFrame:
    """Attach next-season value targets where the next season exists."""
    player_season = player_season.sort_values(["player_id", "season"]).copy()

    next_cols = [
        "season",
        "value_epa_total",
        "value_epa_per_game",
        "value_score",
        "value_score_per_game",
    ]
    for col in next_cols:
        player_season["next_" + col] = player_season.groupby("player_id")[col].shift(-1)

    has_next_consecutive_season = player_season["next_season"].eq(player_season["season"] + 1)
    player_season["next_season_qualifier"] = has_next_consecutive_season.astype(int)

    for col in [
        "next_value_epa_total",
        "next_value_epa_per_game",
        "next_value_score",
        "next_value_score_per_game",
    ]:
        player_season.loc[~has_next_consecutive_season, col] = np.nan

    return player_season


def make_model_pipeline(feature_cols: list[str], model: Any) -> Pipeline:
    """Create a preprocessing + model pipeline for mixed numeric/categorical data."""
    categorical_cols = [col for col in feature_cols if col == "position"]
    numeric_cols = [col for col in feature_cols if col not in categorical_cols]

    numeric_transformer = Pipeline(
        steps=[
            ("imputer", SimpleImputer(strategy="median")),
            ("scaler", StandardScaler()),
        ]
    )
    categorical_transformer = Pipeline(
        steps=[
            ("imputer", SimpleImputer(strategy="most_frequent")),
            ("onehot", OneHotEncoder(handle_unknown="ignore")),
        ]
    )
    preprocessor = ColumnTransformer(
        transformers=[
            ("num", numeric_transformer, numeric_cols),
            ("cat", categorical_transformer, categorical_cols),
        ],
        remainder="drop",
    )

    return Pipeline(
        steps=[
            ("preprocessor", preprocessor),
            ("model", model),
        ]
    )


def make_prediction_pipeline(feature_cols: list[str]) -> Pipeline:
    """Create the tuned next-season value regression pipeline."""
    return make_model_pipeline(
        feature_cols,
        RandomForestRegressor(**TUNED_RANDOM_FOREST_PARAMS),
    )


def make_availability_pipeline(feature_cols: list[str]) -> Pipeline:
    """Create the next-season qualifying-availability classification pipeline."""
    return make_model_pipeline(
        feature_cols,
        RandomForestClassifier(**AVAILABILITY_RANDOM_FOREST_PARAMS),
    )


def evaluate_predictions(y_true: pd.Series, y_pred: np.ndarray) -> dict[str, float]:
    return {
        "mae": mean_absolute_error(y_true, y_pred),
        "rmse": float(np.sqrt(mean_squared_error(y_true, y_pred))),
        "r2": r2_score(y_true, y_pred),
    }


def summarize_residuals_by_position(residuals_df: pd.DataFrame) -> pd.DataFrame:
    """Summarize rolling-validation value-model error by position."""
    if residuals_df.empty:
        return pd.DataFrame()

    return (
        residuals_df
        .groupby("position", as_index=False)
        .agg(
            validation_rows=("player_id", "count"),
            mean_actual_next_value=("next_value_score", "mean"),
            mean_predicted_next_value=("prediction", "mean"),
            bias=("residual", "mean"),
            mae=("abs_residual", "mean"),
            rmse=("residual", lambda s: float(np.sqrt(np.mean(np.square(s))))),
        )
        .sort_values("position")
    )


def summarize_interval_validation(residuals_df: pd.DataFrame) -> pd.DataFrame:
    """Summarize rolling-validation prediction interval coverage."""
    if residuals_df.empty or "interval_contains_actual" not in residuals_df.columns:
        return pd.DataFrame()

    def summarize(grouped: pd.core.groupby.DataFrameGroupBy, segment: str) -> pd.DataFrame:
        summary = grouped.agg(
            validation_rows=("player_id", "count"),
            coverage_rate=("interval_contains_actual", "mean"),
            mean_interval_width=("prediction_interval_width", "mean"),
            mean_uncertainty=("prediction_uncertainty", "mean"),
            mae=("abs_residual", "mean"),
            rmse=("residual", lambda s: float(np.sqrt(np.mean(np.square(s))))),
        ).reset_index()
        summary.insert(0, "segment", segment)
        summary = summary.rename(columns={summary.columns[1]: "segment_value"})
        return summary

    overall = residuals_df.assign(overall="all")
    summary_df = pd.concat(
        [
            summarize(overall.groupby("overall"), "overall"),
            summarize(residuals_df.groupby("position"), "position"),
        ],
        ignore_index=True,
    )
    summary_df["target_coverage"] = PREDICTION_INTERVAL_TARGET_COVERAGE
    summary_df["coverage_gap"] = summary_df["coverage_rate"] - summary_df["target_coverage"]
    return summary_df


def rolling_validation_residuals(
    modeling_df: pd.DataFrame,
    feature_cols: list[str],
    target: str = "next_value_score",
    validation_years: list[int] | None = None,
) -> pd.DataFrame:
    """Run rolling validation to calibrate prediction uncertainty."""
    if validation_years is None:
        validation_years = [2020, 2021, 2022, 2023, 2024]

    records: list[pd.DataFrame] = []
    for valid_year in validation_years:
        fold_train = modeling_df[modeling_df["season"].lt(valid_year)].copy()
        fold_valid = modeling_df[modeling_df["season"].eq(valid_year)].copy()

        if fold_train.empty or fold_valid.empty:
            continue

        pipeline = make_prediction_pipeline(feature_cols)
        pipeline.fit(fold_train[feature_cols], fold_train[target])
        train_pred = pipeline.predict(fold_train[feature_cols])
        fold_train_rmse = float(np.sqrt(mean_squared_error(fold_train[target], train_pred)))
        pred = pipeline.predict(fold_valid[feature_cols])

        model = pipeline.named_steps["model"]
        if hasattr(model, "estimators_"):
            transformed_valid = pipeline.named_steps["preprocessor"].transform(fold_valid[feature_cols])
            tree_predictions = np.column_stack([
                estimator.predict(transformed_valid) for estimator in model.estimators_
            ])
            tree_prediction_std = tree_predictions.std(axis=1)
        else:
            tree_prediction_std = np.zeros(len(fold_valid))
        prediction_uncertainty = np.sqrt(
            np.square(fold_train_rmse) + np.square(tree_prediction_std)
        )
        interval_low = pred - PREDICTION_INTERVAL_MULTIPLIER * prediction_uncertainty
        interval_high = pred + PREDICTION_INTERVAL_MULTIPLIER * prediction_uncertainty

        fold_records = fold_valid[[
            "season",
            "player_id",
            "player_display_name",
            "position",
            target,
        ]].copy()
        fold_records["prediction"] = pred
        fold_records["residual"] = fold_records[target] - fold_records["prediction"]
        fold_records["abs_residual"] = fold_records["residual"].abs()
        fold_records["tree_prediction_std"] = tree_prediction_std
        fold_records["fold_train_rmse"] = fold_train_rmse
        fold_records["prediction_uncertainty"] = prediction_uncertainty
        fold_records["prediction_interval_low"] = interval_low
        fold_records["prediction_interval_high"] = interval_high
        fold_records["prediction_interval_width"] = interval_high - interval_low
        fold_records["interval_contains_actual"] = (
            fold_records[target].between(interval_low, interval_high)
        )
        fold_records["valid_year"] = valid_year
        records.append(fold_records)

    if not records:
        return pd.DataFrame()
    return pd.concat(records, ignore_index=True)


def _positive_class_probability(pipeline: Pipeline, X: pd.DataFrame) -> np.ndarray:
    """Return probability for the positive class, even if a fold has one class."""
    model = pipeline.named_steps["model"]
    probabilities = pipeline.predict_proba(X)
    positive_class_indexes = np.where(model.classes_ == 1)[0]
    if len(positive_class_indexes) == 0:
        return np.zeros(len(X))
    return probabilities[:, positive_class_indexes[0]]


def evaluate_availability_predictions(
    y_true: pd.Series,
    probability: np.ndarray,
) -> dict[str, float]:
    """Evaluate next-season qualifying availability predictions."""
    predicted_class = (probability >= 0.50).astype(int)
    metrics = {
        "accuracy": accuracy_score(y_true, predicted_class),
        "brier_score": brier_score_loss(y_true, probability),
        "positive_rate": float(np.mean(y_true)),
    }
    if y_true.nunique() > 1:
        metrics["roc_auc"] = roc_auc_score(y_true, probability)
    else:
        metrics["roc_auc"] = np.nan
    return metrics


def rolling_availability_validation(
    availability_df: pd.DataFrame,
    feature_cols: list[str],
    target: str = "next_season_qualifier",
    validation_years: list[int] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Run rolling validation for the next-season qualifier model."""
    if validation_years is None:
        validation_years = [2020, 2021, 2022, 2023, 2024]

    records: list[pd.DataFrame] = []
    metrics: list[dict[str, float]] = []

    for valid_year in validation_years:
        fold_train = availability_df[availability_df["season"].lt(valid_year)].copy()
        fold_valid = availability_df[availability_df["season"].eq(valid_year)].copy()

        if fold_train.empty or fold_valid.empty:
            continue

        pipeline = make_availability_pipeline(feature_cols)
        pipeline.fit(fold_train[feature_cols], fold_train[target])
        probability = _positive_class_probability(pipeline, fold_valid[feature_cols])

        fold_records = fold_valid[[
            "season",
            "player_id",
            "player_display_name",
            "position",
            target,
        ]].copy()
        fold_records["predicted_qualifier_probability"] = probability
        fold_records["valid_year"] = valid_year
        records.append(fold_records)

        fold_metrics = evaluate_availability_predictions(fold_valid[target], probability)
        fold_metrics["valid_year"] = valid_year
        metrics.append(fold_metrics)

    records_df = pd.concat(records, ignore_index=True) if records else pd.DataFrame()
    metrics_df = pd.DataFrame(metrics)
    return records_df, metrics_df


def _assign_value_tier(position_percentile: float) -> str:
    if position_percentile >= 0.90:
        return "Elite"
    if position_percentile >= 0.75:
        return "Above Average"
    if position_percentile >= 0.50:
        return "Starter-Level"
    if position_percentile >= 0.25:
        return "Depth/Volatile"
    return "Low Projection"


def _assign_availability_risk(probability: float) -> str:
    if probability >= 0.75:
        return "Low"
    if probability >= 0.55:
        return "Medium"
    return "High"


def _format_confidence_note(row: pd.Series) -> str:
    notes = []
    if row["games_played_2025"] < 8:
        notes.append("limited 2025 sample")
    if row["prediction_uncertainty"] >= row["high_uncertainty_cutoff"]:
        notes.append("wide model interval")
    if row["availability_risk_level"] == "High":
        notes.append("high availability risk")
    if row["confidence_level"] == "High":
        notes.append("strong sample, availability, and model range")
    return "; ".join(notes) if notes else "moderate sample and model range"


def _format_prediction_driver(row: pd.Series) -> str:
    """Create a short, executive-readable explanation for a player projection."""
    drivers = []

    if row["value_score_2025"] >= 1.0:
        drivers.append("strong 2025 value")
    elif row["value_score_2025"] <= -0.75:
        drivers.append("weak 2025 value")

    if pd.notna(row.get("value_score_last2_avg")):
        if row["value_score_last2_avg"] >= 0.75:
            drivers.append("strong recent multi-year value")
        elif row["value_score_last2_avg"] <= -0.50:
            drivers.append("weak recent multi-year value")

    if pd.notna(row.get("value_score_trend_2yr")):
        if row["value_score_trend_2yr"] >= 0.50:
            drivers.append("recent value trend up")
        elif row["value_score_trend_2yr"] <= -0.50:
            drivers.append("recent value trend down")

    if row["games_played_2025"] < 8:
        drivers.append("small 2025 sample")
    elif row.get("games_played_last2_sum", 0) >= 28:
        drivers.append("durable recent sample")

    if row["availability_risk_level"] == "High":
        drivers.append("availability risk")

    if row["prediction_uncertainty"] >= row["high_uncertainty_cutoff"]:
        drivers.append("wide outcome range")

    if not drivers:
        drivers.append("balanced profile")

    return "; ".join(drivers[:4])


def _json_records(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Convert a DataFrame to JSON-safe row dictionaries."""
    return json.loads(df.to_json(orient="records"))


def _clean_sheet_name(name: str) -> str:
    """Keep worksheet names inside Excel's 31-character limit."""
    return name[:31]


def _write_dataframe(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
    startrow: int = 0,
    index: bool = False,
) -> None:
    """Write a DataFrame with a safe empty-frame fallback."""
    output_df = df.copy()
    if output_df.empty:
        output_df = pd.DataFrame({"note": ["No rows available."]})
    output_df.to_excel(
        writer,
        sheet_name=_clean_sheet_name(sheet_name),
        startrow=startrow,
        index=index,
    )


def write_prediction_excel_report(
    outputs: dict[str, Any],
    output_path: str | Path,
) -> Path:
    """Write the recruiter-facing 2026 prediction Excel workbook.

    The workbook is intentionally generated from the same tables used for CSV
    outputs so it can be rebuilt from the command-line pipeline.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to build the Excel report. "
            "Install dependencies with pip install -r requirements.txt."
        ) from exc

    output_path = Path(output_path)
    report_df = outputs["player_predictions"]
    team_summary = outputs["team_summary"]
    position_summary = outputs["position_summary"]
    value_validation = outputs["value_validation_by_position"]
    interval_validation = outputs["interval_validation"]
    availability_metrics = outputs["availability_validation_metrics"]
    data_dictionary = outputs["data_dictionary"]
    model_notes = outputs["model_notes"]

    model_notes_df = pd.DataFrame(
        [
            {
                "setting": key,
                "value": json.dumps(value) if isinstance(value, (dict, list)) else value,
            }
            for key, value in model_notes.items()
        ]
    )

    validation_sections = []
    if not value_validation.empty:
        section = value_validation.copy()
        section.insert(0, "validation_section", "value_model_by_position")
        validation_sections.append(section)
    if not interval_validation.empty:
        section = interval_validation.copy()
        section.insert(0, "validation_section", "prediction_interval_coverage")
        validation_sections.append(section)
    if not availability_metrics.empty:
        section = availability_metrics.copy()
        section.insert(0, "validation_section", "availability_model_by_fold")
        validation_sections.append(section)
    validation_summary = (
        pd.concat(validation_sections, ignore_index=True, sort=False)
        if validation_sections
        else pd.DataFrame({"validation_section": ["No validation rows available."]})
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Dashboard", index=False)
        _write_dataframe(writer, report_df, "Player Predictions")
        _write_dataframe(writer, report_df, "Full Model Data")
        _write_dataframe(writer, team_summary, "Team Summary")
        _write_dataframe(writer, position_summary, "Position Summary")
        _write_dataframe(writer, validation_summary, "Validation Summary")
        _write_dataframe(writer, data_dictionary, "Data Dictionary")
        _write_dataframe(writer, model_notes_df, "Model Notes")

    workbook = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F4E78")
    subtitle_font = Font(size=11, italic=True, color="666666")

    dashboard = workbook["Dashboard"]
    dashboard.delete_rows(1, dashboard.max_row)
    dashboard["A1"] = "2026 NFL Player Value Prediction Report"
    dashboard["A1"].font = title_font
    dashboard["A2"] = "Rebuilt from the reproducible Python pipeline."
    dashboard["A2"].font = subtitle_font

    metric_rows = [
        ("Player projections", len(report_df)),
        ("Average predicted value", round(report_df["predicted_2026_value_score"].mean(), 3)),
        ("Median predicted value", round(report_df["predicted_2026_value_score"].median(), 3)),
        ("High-confidence players", int(report_df["confidence_level"].eq("High").sum())),
        ("Low availability-risk players", int(report_df["availability_risk_level"].eq("Low").sum())),
    ]
    for row_idx, (label, value) in enumerate(metric_rows, start=4):
        dashboard.cell(row_idx, 1, label)
        dashboard.cell(row_idx, 2, value)
        dashboard.cell(row_idx, 1).font = Font(bold=True)

    dashboard["D4"] = "Top 10 Projected Players"
    dashboard["D4"].font = Font(bold=True, color="1F4E78")
    top_cols = [
        "player_display_name",
        "position",
        "primary_team_2025",
        "predicted_2026_value_score",
        "confidence_level",
        "availability_risk_level",
    ]
    top_rows = report_df[top_cols].head(10)
    for col_idx, col in enumerate(top_cols, start=4):
        cell = dashboard.cell(5, col_idx, col)
        cell.fill = header_fill
        cell.font = header_font
    for row_offset, (_, row) in enumerate(top_rows.iterrows(), start=6):
        for col_idx, col in enumerate(top_cols, start=4):
            dashboard.cell(row_offset, col_idx, row[col])

    if not position_summary.empty:
        dashboard["A11"] = "Average Projected Value by Position"
        dashboard["A11"].font = Font(bold=True, color="1F4E78")
        chart_data = position_summary[[
            "position",
            "avg_predicted_2026_value_score",
            "players",
            "avg_qualifying_probability",
        ]].copy()
        for col_idx, col in enumerate(chart_data.columns, start=1):
            cell = dashboard.cell(12, col_idx, col)
            cell.fill = header_fill
            cell.font = header_font
        for row_offset, (_, row) in enumerate(chart_data.iterrows(), start=13):
            dashboard.cell(row_offset, 1, row["position"])
            dashboard.cell(row_offset, 2, row["avg_predicted_2026_value_score"])
            dashboard.cell(row_offset, 3, row["players"])
            dashboard.cell(row_offset, 4, row["avg_qualifying_probability"])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value is not None:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"

        for column_cells in sheet.columns:
            header = column_cells[0].value
            if header is None:
                continue
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[:100]
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_len + 2, 12),
                34,
            )

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)
    return output_path


def build_2026_prediction_tables(
    project_root: Path | None = None,
    save_outputs: bool = True,
) -> dict[str, Any]:
    """Build 2026 prediction tables and optionally save CSV/JSON outputs."""
    if project_root is None:
        project_root = find_project_root()

    processed_dir = project_root / "data" / "processed"
    output_dir = project_root / "outputs" / "tables"
    output_dir.mkdir(parents=True, exist_ok=True)

    skill_seasons_path = processed_dir / "skill_player_seasons_2016_2025.csv"
    value_scores_path = processed_dir / "player_value_scores_2016_2025.csv"
    if skill_seasons_path.exists():
        player_rows = pd.read_csv(skill_seasons_path)
        value_source = "skill_player_seasons_collapsed_before_filter"
    else:
        player_rows = pd.read_csv(value_scores_path)
        value_source = "player_value_scores_existing_file"

    player_season = create_player_season_value_scores(player_rows)
    player_season = add_player_history_features(player_season)
    player_season = create_next_season_targets(player_season)

    feature_cols = [col for col in ENHANCED_FEATURES if col in player_season.columns]
    target = "next_value_score"

    modeling_df = player_season.dropna(subset=[target]).copy()
    training_df = modeling_df[modeling_df["season"].between(2016, 2024)].copy()
    availability_df = player_season[player_season["season"].between(2016, 2024)].copy()
    prediction_input = player_season[player_season["season"].eq(2025)].copy()

    final_model = make_prediction_pipeline(feature_cols)
    final_model.fit(training_df[feature_cols], training_df[target])

    availability_model = make_availability_pipeline(feature_cols)
    availability_model.fit(
        availability_df[feature_cols],
        availability_df["next_season_qualifier"],
    )

    train_pred = final_model.predict(training_df[feature_cols])
    training_metrics = evaluate_predictions(training_df[target], train_pred)

    availability_train_probability = _positive_class_probability(
        availability_model,
        availability_df[feature_cols],
    )
    availability_training_metrics = evaluate_availability_predictions(
        availability_df["next_season_qualifier"],
        availability_train_probability,
    )

    residuals_df = rolling_validation_residuals(modeling_df, feature_cols, target=target)
    availability_validation_df, availability_validation_metrics = rolling_availability_validation(
        availability_df,
        feature_cols,
    )
    if residuals_df.empty:
        residual_rmse = training_metrics["rmse"]
        residual_mae = training_metrics["mae"]
        value_validation_by_position = pd.DataFrame()
        interval_validation = pd.DataFrame()
    else:
        residual_rmse = float(np.sqrt(np.mean(np.square(residuals_df["residual"]))))
        residual_mae = float(residuals_df["abs_residual"].mean())
        value_validation_by_position = summarize_residuals_by_position(residuals_df)
        interval_validation = summarize_interval_validation(residuals_df)

    predicted = prediction_input.copy()
    predicted["predicted_2026_value_score"] = final_model.predict(predicted[feature_cols])
    predicted["predicted_2026_qualifying_probability"] = _positive_class_probability(
        availability_model,
        predicted[feature_cols],
    )

    transformed_2025 = final_model.named_steps["preprocessor"].transform(predicted[feature_cols])
    forest = final_model.named_steps["model"]
    tree_predictions = np.column_stack([
        estimator.predict(transformed_2025) for estimator in forest.estimators_
    ])

    predicted["tree_prediction_std"] = tree_predictions.std(axis=1)
    predicted["tree_prediction_p10"] = np.percentile(tree_predictions, 10, axis=1)
    predicted["tree_prediction_p90"] = np.percentile(tree_predictions, 90, axis=1)
    predicted["prediction_uncertainty"] = np.sqrt(
        np.square(residual_rmse) + np.square(predicted["tree_prediction_std"])
    )
    predicted["prediction_interval_low"] = (
        predicted["predicted_2026_value_score"]
        - PREDICTION_INTERVAL_MULTIPLIER * predicted["prediction_uncertainty"]
    )
    predicted["prediction_interval_high"] = (
        predicted["predicted_2026_value_score"]
        + PREDICTION_INTERVAL_MULTIPLIER * predicted["prediction_uncertainty"]
    )
    predicted["availability_adjusted_2026_value"] = (
        predicted["predicted_2026_value_score"]
        * predicted["predicted_2026_qualifying_probability"]
    )
    predicted["availability_risk_level"] = predicted[
        "predicted_2026_qualifying_probability"
    ].apply(_assign_availability_risk)

    predicted["predicted_2026_overall_percentile"] = predicted["predicted_2026_value_score"].rank(pct=True)
    predicted["predicted_2026_position_percentile"] = (
        predicted.groupby("position")["predicted_2026_value_score"].rank(pct=True)
    )
    predicted["predicted_2026_value_tier"] = predicted["predicted_2026_position_percentile"].apply(_assign_value_tier)

    uncertainty_pct = predicted["prediction_uncertainty"].rank(pct=True)
    sample_score = predicted["games_played"].clip(lower=0, upper=17) / 17
    availability_score = predicted["predicted_2026_qualifying_probability"].clip(0, 1)
    predicted["confidence_score"] = (
        (1 - uncertainty_pct) * 50
        + sample_score * 20
        + availability_score * 30
    ).round(1)
    predicted["confidence_level"] = pd.cut(
        predicted["confidence_score"],
        bins=[-np.inf, 45, 70, np.inf],
        labels=["Low", "Medium", "High"],
    ).astype(str)
    predicted["high_uncertainty_cutoff"] = predicted["prediction_uncertainty"].quantile(0.67)

    report_df = predicted.rename(
        columns={
            "primary_team": "primary_team_2025",
            "teams": "teams_2025",
            "age": "age_2025",
            "years_exp": "years_exp_2025",
            "games_played": "games_played_2025",
            "value_score": "value_score_2025",
            "value_epa_total": "value_epa_total_2025",
            "value_epa_per_game": "value_epa_per_game_2025",
            "yards_per_game": "yards_per_game_2025",
            "tds_per_game": "tds_per_game_2025",
        }
    ).copy()

    report_df["projected_age_2026"] = report_df["age_2025"] + 1
    report_df["projected_years_exp_2026"] = report_df["years_exp_2025"] + 1
    report_df["confidence_note"] = report_df.apply(_format_confidence_note, axis=1)
    report_df["prediction_driver"] = report_df.apply(_format_prediction_driver, axis=1)

    report_cols = [
        "player_id",
        "player_display_name",
        "position",
        "primary_team_2025",
        "teams_2025",
        "games_played_2025",
        "age_2025",
        "projected_age_2026",
        "years_exp_2025",
        "projected_years_exp_2026",
        "draft_number",
        "college",
        "value_score_2025",
        "value_epa_total_2025",
        "value_epa_per_game_2025",
        "yards_per_game_2025",
        "tds_per_game_2025",
        "prior_qualifying_seasons",
        "value_score_prev",
        "value_score_last2_avg",
        "value_score_last3_avg",
        "value_score_trend_2yr",
        "games_played_prev",
        "games_played_last2_sum",
        "predicted_2026_value_score",
        "availability_adjusted_2026_value",
        "predicted_2026_qualifying_probability",
        "predicted_2026_overall_percentile",
        "predicted_2026_position_percentile",
        "predicted_2026_value_tier",
        "availability_risk_level",
        "prediction_interval_low",
        "prediction_interval_high",
        "prediction_uncertainty",
        "confidence_score",
        "confidence_level",
        "confidence_note",
        "prediction_driver",
    ]
    report_cols = [col for col in report_cols if col in report_df.columns]
    report_df = report_df[report_cols].sort_values(
        ["predicted_2026_value_score", "value_score_2025"], ascending=[False, False]
    )

    top_players = report_df.head(30).copy()
    low_confidence = report_df[report_df["confidence_level"].eq("Low")].copy()

    idx_top_team = report_df.groupby("primary_team_2025")["predicted_2026_value_score"].idxmax()
    team_top = report_df.loc[idx_top_team, [
        "primary_team_2025",
        "player_display_name",
        "predicted_2026_value_score",
    ]].rename(
        columns={
            "player_display_name": "top_projected_player",
            "predicted_2026_value_score": "top_projected_value_score",
        }
    )
    team_summary = (
        report_df.groupby("primary_team_2025", as_index=False)
        .agg(
            players=("player_id", "count"),
            avg_predicted_2026_value_score=("predicted_2026_value_score", "mean"),
            avg_availability_adjusted_2026_value=("availability_adjusted_2026_value", "mean"),
            avg_qualifying_probability=("predicted_2026_qualifying_probability", "mean"),
            median_predicted_2026_value_score=("predicted_2026_value_score", "median"),
            avg_confidence_score=("confidence_score", "mean"),
            high_confidence_players=("confidence_level", lambda x: (x == "High").sum()),
            high_availability_risk_players=("availability_risk_level", lambda x: (x == "High").sum()),
            elite_or_above_avg_players=("predicted_2026_value_tier", lambda x: x.isin(["Elite", "Above Average"]).sum()),
        )
        .merge(team_top, on="primary_team_2025", how="left")
        .sort_values("avg_predicted_2026_value_score", ascending=False)
    )

    idx_top_position = report_df.groupby("position")["predicted_2026_value_score"].idxmax()
    position_top = report_df.loc[idx_top_position, [
        "position",
        "player_display_name",
        "predicted_2026_value_score",
    ]].rename(
        columns={
            "player_display_name": "top_projected_player",
            "predicted_2026_value_score": "top_projected_value_score",
        }
    )
    position_summary = (
        report_df.groupby("position", as_index=False)
        .agg(
            players=("player_id", "count"),
            avg_predicted_2026_value_score=("predicted_2026_value_score", "mean"),
            avg_availability_adjusted_2026_value=("availability_adjusted_2026_value", "mean"),
            avg_qualifying_probability=("predicted_2026_qualifying_probability", "mean"),
            median_predicted_2026_value_score=("predicted_2026_value_score", "median"),
            avg_confidence_score=("confidence_score", "mean"),
            high_confidence_players=("confidence_level", lambda x: (x == "High").sum()),
            high_availability_risk_players=("availability_risk_level", lambda x: (x == "High").sum()),
            low_confidence_players=("confidence_level", lambda x: (x == "Low").sum()),
        )
        .merge(position_top, on="position", how="left")
        .sort_values("avg_predicted_2026_value_score", ascending=False)
    )

    model_notes = {
        "report_name": "2026 NFL Player Value Prediction Report",
        "prediction_target": "2026 position-adjusted value_score",
        "training_rows": int(len(training_df)),
        "prediction_rows_2025": int(len(report_df)),
        "training_seasons": "2016-2024 current-season rows with known next-season outcomes",
        "prediction_input_season": 2025,
        "value_model": "Enhanced-history tuned RandomForestRegressor",
        "availability_model": "RandomForestClassifier",
        "feature_set": "raw_production_plus_multi_year_history",
        "value_source": value_source,
        "minimum_games_for_value_score": MIN_VALUE_GAMES,
        "features": feature_cols,
        "value_model_params": TUNED_RANDOM_FOREST_PARAMS,
        "availability_model_params": AVAILABILITY_RANDOM_FOREST_PARAMS,
        "training_rmse": float(training_metrics["rmse"]),
        "training_mae": float(training_metrics["mae"]),
        "training_r2": float(training_metrics["r2"]),
        "rolling_validation_residual_rmse": residual_rmse,
        "rolling_validation_abs_residual_mae": residual_mae,
        "value_rolling_validation_by_position": _json_records(value_validation_by_position),
        "prediction_interval_multiplier": PREDICTION_INTERVAL_MULTIPLIER,
        "prediction_interval_target_coverage": PREDICTION_INTERVAL_TARGET_COVERAGE,
        "prediction_interval_validation": _json_records(interval_validation),
        "availability_training_accuracy": float(availability_training_metrics["accuracy"]),
        "availability_training_roc_auc": float(availability_training_metrics["roc_auc"]),
        "availability_training_brier_score": float(availability_training_metrics["brier_score"]),
        "availability_rolling_validation_mean_roc_auc": (
            float(availability_validation_metrics["roc_auc"].mean())
            if not availability_validation_metrics.empty
            else None
        ),
        "availability_rolling_validation_mean_brier_score": (
            float(availability_validation_metrics["brier_score"].mean())
            if not availability_validation_metrics.empty
            else None
        ),
        "confidence_method": (
            "Confidence combines calibrated rolling-validation error, tree-level model disagreement, "
            "2025 games played, and predicted next-season qualifying probability. "
            "It is a practical uncertainty label, not a guarantee."
        ),
        "availability_method": (
            "The availability model estimates the probability that a player will have a "
            "qualifying next-season row, which partially addresses survivorship bias."
        ),
        "team_note": "primary_team_2025 is the team with the largest 2025 games sample; teams_2025 lists all teams in the player-season data.",
    }

    data_dictionary = pd.DataFrame([
        {"column": "predicted_2026_value_score", "definition": "Model projection for next-season position-adjusted value score."},
        {"column": "availability_adjusted_2026_value", "definition": "Predicted value score multiplied by the model-estimated probability of a qualifying 2026 season."},
        {"column": "predicted_2026_qualifying_probability", "definition": "Estimated probability that the player has a qualifying next-season row."},
        {"column": "availability_risk_level", "definition": "Low, Medium, or High availability risk derived from qualifying probability."},
        {"column": "predicted_2026_position_percentile", "definition": "Projected percentile within the player's listed position."},
        {"column": "prediction_interval_low/high", "definition": "Approximate central 80% interval around the prediction using model disagreement and rolling-validation error; coverage is checked on rolling-validation folds."},
        {"column": "confidence_score", "definition": "0-100 practical confidence score based on model uncertainty and 2025 sample size."},
        {"column": "confidence_level", "definition": "High, Medium, or Low label derived from confidence_score."},
        {"column": "prediction_driver", "definition": "Short plain-English summary of the main signals and risks behind the projection."},
        {"column": "value_score_last2_avg", "definition": "Average value score over the player's prior two qualifying seasons."},
        {"column": "value_score_trend_2yr", "definition": "Previous-season value score minus the prior-two-season average; positive values indicate recent improvement."},
        {"column": "games_played_last2_sum", "definition": "Total qualifying games played over the prior two seasons."},
        {"column": "value_score_2025", "definition": "Actual 2025 standardized total EPA value score within season-position group."},
        {"column": "value_epa_total_2025", "definition": "Raw 2025 EPA used for value: QB EPA for QBs, scrimmage EPA for RB/WR/TE."},
        {"column": "primary_team_2025", "definition": "Team with the largest 2025 games sample in the player-season data."},
        {"column": "teams_2025", "definition": "All teams appearing for that player in 2025, slash-separated for multi-team seasons."},
    ])

    outputs = {
        "player_predictions": report_df,
        "team_summary": team_summary,
        "position_summary": position_summary,
        "top_players": top_players,
        "low_confidence": low_confidence,
        "data_dictionary": data_dictionary,
        "model_notes": model_notes,
        "residuals": residuals_df,
        "availability_validation": availability_validation_df,
        "availability_validation_metrics": availability_validation_metrics,
        "value_validation_by_position": value_validation_by_position,
        "interval_validation": interval_validation,
        "output_dir": output_dir,
    }

    if save_outputs:
        report_df.to_csv(
            output_dir / "2026_player_value_predictions.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        team_summary.to_csv(
            output_dir / "2026_team_summary.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        position_summary.to_csv(
            output_dir / "2026_position_summary.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        top_players.to_csv(
            output_dir / "2026_top_predicted_players.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        low_confidence.to_csv(
            output_dir / "2026_low_confidence_predictions.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        data_dictionary.to_csv(
            output_dir / "2026_prediction_data_dictionary.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        residuals_df.to_csv(
            output_dir / "2026_prediction_validation_residuals.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        value_validation_by_position.to_csv(
            output_dir / "2026_value_validation_by_position.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        interval_validation.to_csv(
            output_dir / "2026_prediction_interval_validation.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        availability_validation_df.to_csv(
            output_dir / "2026_availability_validation_predictions.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        availability_validation_metrics.to_csv(
            output_dir / "2026_availability_validation_metrics.csv",
            index=False,
            float_format=CSV_FLOAT_FORMAT,
        )
        with (output_dir / "2026_prediction_model_notes.json").open("w") as f:
            json.dump(model_notes, f, indent=2)

        report_payload = {
            "player_predictions": _json_records(report_df),
            "team_summary": _json_records(team_summary),
            "position_summary": _json_records(position_summary),
            "top_players": _json_records(top_players),
            "low_confidence": _json_records(low_confidence),
            "data_dictionary": _json_records(data_dictionary),
            "value_validation_by_position": _json_records(value_validation_by_position),
            "interval_validation": _json_records(interval_validation),
            "availability_validation_metrics": _json_records(availability_validation_metrics),
            "model_notes": model_notes,
        }
        with (output_dir / "2026_prediction_report_tables.json").open("w") as f:
            json.dump(report_payload, f, indent=2)

        write_prediction_excel_report(
            outputs,
            output_dir / "2026_player_value_predictions.xlsx",
        )

    return outputs
